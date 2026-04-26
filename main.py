import os
import json
import csv
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from typing import List, Optional
import threading
from dataclasses import dataclass, asdict
import time
import webbrowser

# Optional imports - app works without them
try:
    import requests
    from bs4 import BeautifulSoup
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    HAS_SELENIUM = True
except ImportError:
    HAS_SELENIUM = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class ScrapedItem:
    category: str
    title: str
    link: str
    timestamp: str
    company: Optional[str] = None
    location: Optional[str] = None
    price: Optional[str] = None
    rating: Optional[str] = None
    description: Optional[str] = None
    source: Optional[str] = None
    date_published: Optional[str] = None


class WebScraper:
    HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36'
    }

    URLS = {
        'jobs': 'https://news.ycombinator.com/jobs',
        'products': 'http://books.toscrape.com/',
        'news': 'https://news.ycombinator.com/'
    }

    def __init__(self):
        self.session = None
        if HAS_REQUESTS:
            self.session = requests.Session()
            self.session.headers.update(self.HEADERS)
        self.driver = None
        self.selenium_available = self._check_selenium()

    def _check_selenium(self) -> bool:
        if not HAS_SELENIUM:
            return False
        try:
            opts = Options()
            opts.add_argument('--headless')
            opts.add_argument('--no-sandbox')
            d = webdriver.Chrome(options=opts)
            d.quit()
            return True
        except Exception:
            return False

    def fetch_page(self, url: str):
        if not HAS_REQUESTS or not self.session:
            raise ConnectionError("requests library not installed")
        resp = self.session.get(url, timeout=10)
        resp.raise_for_status()
        return BeautifulSoup(resp.content, 'html.parser')

    def scrape_jobs(self) -> List[ScrapedItem]:
        jobs = []
        soup = self.fetch_page(self.URLS['jobs'])
        rows = soup.find_all('tr', class_='athing')[:20]
        for row in rows:
            try:
                a = row.find('span', class_='titleline')
                if not a:
                    continue
                link = a.find('a')
                title = link.get_text(strip=True) if link else 'N/A'
                href = link.get('href', '') if link else ''
                if href.startswith('item?'):
                    href = 'https://news.ycombinator.com/' + href

                company = 'N/A'
                if ' at ' in title:
                    company = title.split(' at ', 1)[1].strip()

                jobs.append(ScrapedItem(
                    category='Jobs', title=title, company=company,
                    location='Remote / See post', link=href,
                    timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
            except Exception:
                continue
        return jobs

    def scrape_products(self) -> List[ScrapedItem]:
        products = []
        soup = self.fetch_page(self.URLS['products'])
        items = soup.find_all('article', class_='product_pod')[:20]
        for item in items:
            try:
                title_a = item.find('h3').find('a')
                price = item.find('p', class_='price_color')
                rating = item.find('p', class_='star-rating')
                rating_val = 'N/A'
                if rating and rating.get('class'):
                    rating_val = [c for c in rating.get('class') if c != 'star-rating'][0]

                products.append(ScrapedItem(
                    category='Products',
                    title=title_a.get('title', 'N/A'),
                    price=price.get_text(strip=True) if price else 'N/A',
                    rating=rating_val,
                    link=self.URLS['products'] + title_a.get('href', ''),
                    timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
            except Exception:
                continue
        return products

    def scrape_news(self) -> List[ScrapedItem]:
        news = []
        soup = self.fetch_page(self.URLS['news'])
        stories = soup.find_all('tr', class_='athing')[:20]
        for story in stories:
            try:
                title_a = story.find('span', class_='titleline').find('a')
                sub = story.find_next_sibling('tr')
                age = sub.find('span', class_='age').get_text(strip=True) if sub else 'Unknown'
                score = sub.find('span', class_='score').get_text(strip=True) if sub else 'N/A'

                news.append(ScrapedItem(
                    category='News', title=title_a.get_text(strip=True),
                    link=title_a.get('href', ''), source='Hacker News',
                    description=f"{score} | Posted {age}", date_published=age,
                    timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
            except Exception:
                continue
        return news

    def generate_demo_data(self, category: str) -> List[ScrapedItem]:
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if category == 'Jobs':
            return [
                ScrapedItem('Jobs', 'Senior Python Developer', 'https://example.com/job1', ts, company='Google', location='Remote'),
                ScrapedItem('Jobs', 'Backend Engineer at Stripe', 'https://example.com/job2', ts, company='Stripe', location='San Francisco'),
                ScrapedItem('Jobs', 'ML Engineer', 'https://example.com/job3', ts, company='OpenAI', location='New York'),
                ScrapedItem('Jobs', 'Full Stack Developer', 'https://example.com/job4', ts, company='Meta', location='London'),
                ScrapedItem('Jobs', 'DevOps Engineer', 'https://example.com/job5', ts, company='Netflix', location='Remote'),
            ]
        elif category == 'Products':
            return [
                ScrapedItem('Products', 'A Light in the Attic', 'https://example.com/p1', ts, price='£51.77', rating='Three'),
                ScrapedItem('Products', 'Tipping the Velvet', 'https://example.com/p2', ts, price='£53.74', rating='One'),
                ScrapedItem('Products', 'Soumission', 'https://example.com/p3', ts, price='£50.10', rating='One'),
                ScrapedItem('Products', 'Sharp Objects', 'https://example.com/p4', ts, price='£47.82', rating='Four'),
                ScrapedItem('Products', 'Sapiens', 'https://example.com/p5', ts, price='£54.23', rating='Five'),
            ]
        else:
            return [
                ScrapedItem('News', 'Show HN: I built a web scraper in Python', 'https://example.com/n1', ts, source='Hacker News', description='128 points | Posted 2 hours ago', date_published='2h ago'),
                ScrapedItem('News', 'Python 4.0 announced', 'https://example.com/n2', ts, source='Hacker News', description='342 points | Posted 5 hours ago', date_published='5h ago'),
                ScrapedItem('News', 'Why I switched from Node to Python', 'https://example.com/n3', ts, source='Hacker News', description='89 points | Posted 8 hours ago', date_published='8h ago'),
                ScrapedItem('News', 'The future of AI agents', 'https://example.com/n4', ts, source='Hacker News', description='512 points | Posted 1 hour ago', date_published='1h ago'),
                ScrapedItem('News', 'Docker best practices 2026', 'https://example.com/n5', ts, source='Hacker News', description='201 points | Posted 3 hours ago', date_published='3h ago'),
            ]


class ScraperGUI:
    COLORS = {
        'primary': '#1e3a8a', 'secondary': '#3b82f6', 'accent': '#60a5fa',
        'success': '#10b981', 'warning': '#f59e0b', 'danger': '#ef4444',
        'bg_light': '#f8fafc', 'text_dark': '#1e293b', 'text_light': '#ffffff',
        'hover': 'dimgrey'
    }

    def __init__(self, root):
        self.root = root
        self.root.title("Professional Web Scraper Pro")
        self.root.geometry("1250x800")
        self.root.configure(bg=self.COLORS['bg_light'])
        self.scraper = WebScraper()
        self.data: List[ScrapedItem] = []
        self.filtered_data: List[ScrapedItem] = []

        # Auto-refresh state
        self.auto_refresh_on = False
        self.refresh_timer_id = None
        self.refresh_interval_sec = 60

        self.setup_ui()

    def setup_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg=self.COLORS['primary'], height=80)
        hdr.pack(fill='x')
        tk.Label(hdr, text="Professional Web Scraper", font=('Segoe UI', 24, 'bold'),
                bg=self.COLORS['primary'], fg='white').pack(pady=15)
        tk.Label(hdr, text="Extract Jobs • Products • News | Auto-Refresh Supported",
                bg=self.COLORS['primary'], fg=self.COLORS['accent']).pack()

        # Controls
        ctrl = tk.Frame(self.root, bg=self.COLORS['bg_light'], padx=15, pady=10)
        ctrl.pack(fill='x')

        # Row 1: Category | Scrape | Demo | Auto-Refresh | Interval | Apply | Status
        row1 = tk.Frame(ctrl, bg=self.COLORS['bg_light'])
        row1.pack(fill='x')

        tk.Label(row1, text="Category:", bg=self.COLORS['bg_light'], font=('Segoe UI', 10, 'bold')).pack(side='left')
        self.cat_var = tk.StringVar(value='Jobs')
        ttk.Combobox(row1, textvariable=self.cat_var, values=['Jobs', 'Products', 'News'],
                    state='readonly', width=12).pack(side='left', padx=5)

        tk.Button(row1, text="Scrape Now", command=self.scrape_data,
                 bg=self.COLORS['success'], fg='white', font=('Segoe UI', 10, 'bold'),
                 relief='flat', padx=15, pady=4, cursor='hand2').pack(side='left', padx=5)

        tk.Button(row1, text="Load Demo Data", command=self.load_demo,
                 bg=self.COLORS['warning'], fg='white', font=('Segoe UI', 10, 'bold'),
                 relief='flat', padx=15, pady=4, cursor='hand2').pack(side='left', padx=5)

        # Auto-Refresh merged into row 1
        tk.Frame(row1, bg=self.COLORS['bg_light'], width=20).pack(side='left')

        self.auto_var = tk.BooleanVar(value=False)
        self.auto_cb = tk.Checkbutton(
            row1, text="Auto-Refresh", variable=self.auto_var,
            command=self.toggle_auto_refresh, bg=self.COLORS['bg_light'],
            font=('Segoe UI', 10, 'bold'), fg=self.COLORS['text_dark'],
            selectcolor=self.COLORS['bg_light']
        )
        self.auto_cb.pack(side='left', padx=(5, 2))

        tk.Label(row1, text="Interval (sec):", bg=self.COLORS['bg_light'],
                font=('Segoe UI', 10)).pack(side='left', padx=(5, 2))

        self.interval_var = tk.StringVar(value="60")
        self.interval_entry = tk.Entry(row1, textvariable=self.interval_var, width=6,
                                      font=('Segoe UI', 10), justify='center')
        self.interval_entry.pack(side='left')
        self.interval_entry.bind('<Return>', lambda e: self.update_interval())

        tk.Button(row1, text="Apply", command=self.update_interval,
                 bg=self.COLORS['secondary'], fg='white', relief='flat',
                 padx=10, font=('Segoe UI', 9)).pack(side='left', padx=4)

        self.refresh_status = tk.Label(
            row1, text="OFF", bg=self.COLORS['bg_light'],
            fg=self.COLORS['danger'], font=('Segoe UI', 10, 'bold')
        )
        self.refresh_status.pack(side='left', padx=(8, 0))

        # Row 2: Save buttons + Search
        row2 = tk.Frame(ctrl, bg=self.COLORS['bg_light'])
        row2.pack(fill='x', pady=(10, 0))

        tk.Button(row2, text="Save JSON", command=lambda: self.save('json'),
                 bg=self.COLORS['secondary'], fg='white', relief='flat', padx=10).pack(side='left', padx=2)
        tk.Button(row2, text="Save CSV", command=lambda: self.save('csv'),
                 bg=self.COLORS['secondary'], fg='white', relief='flat', padx=10).pack(side='left', padx=2)
        tk.Button(row2, text="Save Excel", command=self.save_excel,
                 bg=self.COLORS['success'], fg='white', relief='flat', padx=10).pack(side='left', padx=2)
        tk.Button(row2, text="Clear", command=self.clear,
                 bg=self.COLORS['danger'], fg='white', relief='flat', padx=10).pack(side='left', padx=2)

        tk.Label(row2, text="Search:", bg=self.COLORS['bg_light'], font=('Segoe UI', 10, 'bold')).pack(side='left', padx=(20, 5))
        self.search_var = tk.StringVar()
        tk.Entry(row2, textvariable=self.search_var, width=35, font=('Segoe UI', 10)).pack(side='left')
        tk.Button(row2, text="Filter", command=self.filter_data,
                 bg=self.COLORS['accent'], fg='white', relief='flat', padx=12).pack(side='left', padx=3)
        tk.Button(row2, text="Reset", command=self.reset_filter,
                 bg=self.COLORS['hover'], fg='white', relief='flat', padx=12).pack(side='left', padx=3)

        self.count_lbl = tk.Label(row2, text="Items: 0", bg=self.COLORS['bg_light'],
                                  fg=self.COLORS['secondary'], font=('Segoe UI', 10, 'bold'))
        self.count_lbl.pack(side='right', padx=10)

        # Treeview (expanded area)
        frm = tk.Frame(self.root, bg=self.COLORS['bg_light'])
        frm.pack(fill='both', expand=True, padx=15, pady=8)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Custom.Treeview', rowheight=30, font=('Segoe UI', 9))
        style.configure('Custom.Treeview.Heading', background=self.COLORS['primary'],
                       foreground='white', font=('Segoe UI', 10, 'bold'))

        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        self.tree = ttk.Treeview(frm, yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                style='Custom.Treeview')
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')
        self.tree.pack(fill='both', expand=True)

        self.tree['columns'] = ('Category', 'Title', 'Details', 'Link')
        self.tree.column('#0', width=0, stretch=False)
        self.tree.column('Category', width=80, anchor='center')
        self.tree.column('Title', width=380)
        self.tree.column('Details', width=300)
        self.tree.column('Link', width=350)
        for c in self.tree['columns']:
            self.tree.heading(c, text=c)

        # Link styling and interactions
        self.tree.tag_configure('link', foreground=self.COLORS['secondary'])
        self.tree.bind('<Double-Button-1>', self.open_link)

        # HAND CURSOR ON LINK HOVER
        self.tree.bind('<Motion>', self.on_mouse_motion)

        # Status bar
        self.status_var = tk.StringVar(value="Ready | Hover over Link column to see hand cursor")
        tk.Frame(self.root, bg=self.COLORS['primary'], height=35).pack(side='bottom', fill='x')
        tk.Label(self.root, textvariable=self.status_var, bg=self.COLORS['primary'],
                fg='white', anchor='w', padx=15).pack(side='bottom', fill='x')

    # ------------------------------------------------------------------
    # HAND CURSOR ON LINK COLUMN HOVER
    # ------------------------------------------------------------------
    def on_mouse_motion(self, event):
        """Change cursor to hand2 when hovering over the Link column"""
        region = self.tree.identify('region', event.x, event.y)
        column = self.tree.identify_column(event.x)

        if region == 'cell' and column == '#4':
            self.tree.config(cursor='hand2')
        else:
            self.tree.config(cursor='')

    # ------------------------------------------------------------------
    # Auto-Refresh Logic
    # ------------------------------------------------------------------
    def update_interval(self):
        try:
            val = int(self.interval_var.get())
            if val < 5:
                val = 5
                self.interval_var.set("5")
            self.refresh_interval_sec = val
            self.status_var.set(f"Refresh interval set to {val} seconds")
        except ValueError:
            messagebox.showwarning("Invalid", "Enter a number (seconds)")
            self.interval_var.set("60")

    def toggle_auto_refresh(self):
        if self.auto_var.get():
            self.update_interval()
            self.auto_refresh_on = True
            self.refresh_status.config(text="ON", fg=self.COLORS['success'])
            self.status_var.set(f"Auto-refresh started: every {self.refresh_interval_sec} seconds")
            self._schedule_refresh()
        else:
            self.auto_refresh_on = False
            self.refresh_status.config(text="OFF", fg=self.COLORS['danger'])
            self.status_var.set("Auto-refresh stopped")
            if self.refresh_timer_id:
                self.root.after_cancel(self.refresh_timer_id)
                self.refresh_timer_id = None

    def _schedule_refresh(self):
        if not self.auto_refresh_on:
            return
        self.refresh_timer_id = self.root.after(self.refresh_interval_sec * 1000, self._do_refresh)

    def _do_refresh(self):
        if not self.auto_refresh_on:
            return
        self.status_var.set(f"Auto-refresh triggered: scraping {self.cat_var.get()}...")
        self._run_scrape_thread(auto_mode=True)

    # ------------------------------------------------------------------
    # Scraping & Display
    # ------------------------------------------------------------------
    def open_link(self, event):
        item = self.tree.selection()
        if not item:
            return
        vals = self.tree.item(item[0], 'values')
        if len(vals) >= 4:
            link = vals[3]
            if link.startswith('http'):
                webbrowser.open(link)

    def load_demo(self):
        cat = self.cat_var.get()
        self.data = self.scraper.generate_demo_data(cat)
        self.filtered_data = self.data.copy()
        self.update_display()
        self.status_var.set(f"Loaded {len(self.data)} demo {cat}")

    def scrape_data(self):
        self.status_var.set(f"Scraping {self.cat_var.get()}...")
        self.root.config(cursor="wait")
        self._run_scrape_thread(auto_mode=False)

    def _run_scrape_thread(self, auto_mode: bool):
        def worker():
            try:
                cat = self.cat_var.get()
                if cat == 'Jobs':
                    res = self.scraper.scrape_jobs()
                elif cat == 'Products':
                    res = self.scraper.scrape_products()
                else:
                    res = self.scraper.scrape_news()

                self.data = res
                self.filtered_data = res.copy()
                self.root.after(0, self.update_display)

                if auto_mode:
                    self.root.after(0, lambda: self.status_var.set(
                        f"Auto-refreshed: {len(res)} {cat} | Next in {self.refresh_interval_sec}s"
                    ))
                    self.root.after(0, self._schedule_refresh)
                else:
                    self.root.after(0, lambda: self.status_var.set(f"Scraped {len(res)} {cat}"))

            except Exception as e:
                msg = str(e)
                if "getaddrinfo" in msg or "NameResolutionError" in msg:
                    hint = "No internet. Use 'Load Demo Data'."
                elif "driver" in msg.lower():
                    hint = "ChromeDriver missing. Use 'Load Demo Data'."
                else:
                    hint = msg[:60]
                self.root.after(0, lambda: self.status_var.set(f"Error: {hint}"))
                if auto_mode:
                    self.root.after(0, self._schedule_refresh)
            finally:
                self.root.after(0, lambda: self.root.config(cursor=""))

        threading.Thread(target=worker, daemon=True).start()

    def filter_data(self):
        q = self.search_var.get().lower().strip()
        if not q:
            self.filtered_data = self.data.copy()
        else:
            self.filtered_data = [
                d for d in self.data
                if q in d.title.lower() or
                   (d.company and q in d.company.lower()) or
                   (d.description and q in d.description.lower())
            ]
        self.update_display()
        self.status_var.set(f"Filter matched {len(self.filtered_data)} items")

    def reset_filter(self):
        self.search_var.set("")
        self.filtered_data = self.data.copy()
        self.update_display()

    def update_display(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        display = self.filtered_data if self.search_var.get() else self.data
        for item in display:
            details = []
            if item.company:
                details.append(f"Company: {item.company}")
            if item.location:
                details.append(f"Loc: {item.location}")
            if item.price:
                details.append(f"Price: {item.price}")
            if item.rating:
                details.append(f"Rating: {item.rating}")
            if item.source:
                details.append(f"Src: {item.source}")
            if item.date_published:
                details.append(f"Date: {item.date_published}")

            self.tree.insert('', 'end', values=(
                item.category,
                item.title[:65] + '...' if len(item.title) > 65 else item.title,
                ' | '.join(details),
                item.link[:55] + '...' if len(item.link) > 55 else item.link
            ), tags=('link',))
        self.count_lbl.config(text=f"Items: {len(display)}")

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    def save(self, fmt: str):
        if not self.data:
            messagebox.showwarning("No Data", "Scrape or load demo data first.")
            return
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = filedialog.asksaveasfilename(
            defaultextension=f".{fmt}", initialfile=f"data_{ts}.{fmt}",
            filetypes=[(fmt.upper(), f"*.{fmt}")]
        )
        if not path:
            return
        try:
            if fmt == 'json':
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump([asdict(i) for i in self.data], f, indent=2, ensure_ascii=False)
            else:
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    w = csv.DictWriter(f, fieldnames=asdict(self.data[0]).keys())
                    w.writeheader()
                    w.writerows([asdict(i) for i in self.data])
            self.status_var.set(f"Saved to {path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def save_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showwarning("Missing", "Install openpyxl: pip install openpyxl")
            return
        if not self.data:
            messagebox.showwarning("No Data", "Scrape or load demo data first.")
            return
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile=f"data_{ts}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            headers = list(asdict(self.data[0]).keys())
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            for item in self.data:
                ws.append(list(asdict(item).values()))
            wb.save(path)
            self.status_var.set(f"Excel saved: {path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def clear(self):
        if messagebox.askyesno("Clear", "Remove all data?"):
            self.data.clear()
            self.filtered_data.clear()
            self.update_display()
            self.status_var.set("Cleared")


def main():
    root = tk.Tk()
    ScraperGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
